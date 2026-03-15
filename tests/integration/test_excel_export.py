from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from causal_app.export.excel import export_pipeline_results
from causal_app.schemas.contracts import EXPORT_SHEETS


def test_export_pipeline_results_creates_expected_workbook(tmp_path):
    export_path = tmp_path / "demo.xlsx"
    diagnostics = {
        "input_rows": 3,
        "accepted_rows": 2,
        "rejected_rows": 1,
        "validation_reject_reason_counts": {"duplicate_customerid": 1},
        "stage_results": [
            {
                "stage_name": "read_input",
                "status": "completed",
                "input_rows": 3,
                "output_rows": 3,
                "duration_seconds": 0.01,
            }
        ],
    }
    recommendations = pd.DataFrame(
        [
            {
                "source_row_number": 2,
                "CustomerId": 1001,
                "Surname": "Alpha",
                "Geography": "Germany",
                "Age": 45,
                "Balance": 10000.0,
                "NumOfProducts": 1,
                "churn_probability": 0.91,
                "assigned_cluster": 1,
                "recommended_treatment": "Engage & Elevate",
                "estimated_post_churn": 0.85,
                "expected_absolute_change": -0.06,
                "policy_scope": "cluster_specific",
                "policy_sample_size": 42,
            }
        ]
    )
    rejected_rows = pd.DataFrame(
        [
            {
                "_source_row_number": 7,
                "CustomerId": 9999,
                "Surname": "Reject",
                "Geography": "France",
                "reject_reasons": "duplicate_customerid",
            }
        ]
    )

    export_pipeline_results(
        export_path=export_path,
        diagnostics=diagnostics,
        recommendations=recommendations,
        rejected_rows=rejected_rows,
        policy_options=pd.DataFrame(),
        run_id="demo-run",
        input_path=Path("input.csv"),
    )

    workbook = load_workbook(export_path, read_only=True, data_only=True)

    assert tuple(workbook.sheetnames) == EXPORT_SHEETS
    assert workbook["Customer_Action_List"]["Q1"].value == "priority_band"
    assert workbook["Reject_Report"]["K1"].value == "reject_reason_detail"
    assert workbook["Summary"]["A2"].value == "Run"
