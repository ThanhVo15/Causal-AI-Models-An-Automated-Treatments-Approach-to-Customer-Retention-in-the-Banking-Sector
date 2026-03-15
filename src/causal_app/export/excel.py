from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

from causal_app.export.business_output import (
    build_customer_action_list,
    build_field_definitions_sheet,
    build_reject_report,
    build_run_metadata_sheet,
    build_summary_sheet,
)


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _format_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)


def export_pipeline_results(
    export_path: Path,
    diagnostics: dict,
    recommendations: pd.DataFrame,
    rejected_rows: pd.DataFrame,
    policy_options: pd.DataFrame,
    *,
    run_id: str,
    input_path: Path,
) -> Path:
    export_path.parent.mkdir(parents=True, exist_ok=True)

    customer_action_list = build_customer_action_list(recommendations=recommendations, run_id=run_id)
    reject_report = build_reject_report(rejected_rows=rejected_rows, run_id=run_id, input_path=input_path)
    summary_sheet = build_summary_sheet(
        diagnostics=diagnostics,
        action_list=customer_action_list,
        reject_report=reject_report,
        run_id=run_id,
        input_path=input_path,
    )
    run_metadata_sheet = build_run_metadata_sheet(
        diagnostics=diagnostics,
        run_id=run_id,
        input_path=input_path,
        export_path=export_path,
    )
    field_definitions_sheet = build_field_definitions_sheet()

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        summary_sheet.to_excel(writer, sheet_name="Summary", index=False)
        customer_action_list.to_excel(writer, sheet_name="Customer_Action_List", index=False)
        reject_report.to_excel(writer, sheet_name="Reject_Report", index=False)
        run_metadata_sheet.to_excel(writer, sheet_name="Run_Metadata", index=False)
        field_definitions_sheet.to_excel(writer, sheet_name="Field_Definitions", index=False)

        for worksheet in writer.book.worksheets:
            _format_worksheet(worksheet)

    return export_path
